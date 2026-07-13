# SPDX-License-Identifier: GPL-3.0-or-later
"""Generate ``tracking.xlsx`` from the database — one sheet per pack category.

Only confirmed-missing packs are drawn red (Standard numeric gaps, or anything a
loaded osu! API reference validates). Everything else is listed plainly. Adds a
Summary sheet and an Artists sheet.
"""
from __future__ import annotations

from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from . import gaps, osu_api
from .gaps import _present_row

_RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_RED_FONT = Font(color="9C0006")
_HEADER_FONT = Font(bold=True)
_HEADER_FILL = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

_COLS = ["Series", "Number / Season", "Code", "Title", "Mode", "Tracks", "Extracted At"]
_WIDTHS = [10, 16, 12, 44, 12, 8, 20]


def _header(ws, cols):
    for c, name in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    ws.freeze_panes = "A2"


def _widths(ws, widths):
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w


def _row_key(row):
    if row.series == "R" or row.year is not None:
        return " ".join(str(x) for x in (row.year, row.season) if x is not None)
    return row.number


def _category_sheet(wb, category, rows):
    ws = wb.create_sheet(title=category[:31])
    _header(ws, _COLS)
    _widths(ws, _WIDTHS)
    r = 2
    for row in rows:
        values = [row.series or "", _row_key(row), row.code or "", row.title or "",
                  row.mode or "", row.track_count if row.present else "",
                  row.extracted_at or ""]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value="" if v is None else v)
            if not row.present:
                cell.fill = _RED_FILL
                cell.font = _RED_FONT
        r += 1


def _artists_sheet(wb, db):
    ws = wb.create_sheet(title="Artists")
    for c, name in enumerate(["Artist", "Songs"], start=1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    ws.freeze_panes = "A2"
    for i, a in enumerate(db.artists_by_count(descending=True), start=2):
        ws.cell(row=i, column=1, value=a["artist"])
        ws.cell(row=i, column=2, value=a["song_count"])
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 10


def _summary_sheet(wb, db, numbered_missing):
    ws = wb.create_sheet(title="Summary")
    counts = db.counts()
    ws["A1"] = "osu! Archive Manager — Summary"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"], ws["B3"] = "Packs", counts["packs"]
    ws["A4"], ws["B4"] = "Tracks", counts["tracks"]
    ws["A5"], ws["B5"] = "In Library", counts["in_library"]
    ws["A7"] = "Series"; ws["B7"] = "Confirmed missing"
    for c in ("A7", "B7"):
        ws[c].font = _HEADER_FONT
        ws[c].fill = _HEADER_FILL
    r = 8
    for series, miss in sorted(numbered_missing.items()):
        ws.cell(row=r, column=1, value=series)
        cell = ws.cell(row=r, column=2, value=", ".join(map(str, miss)))
        cell.fill = _RED_FILL
        cell.font = _RED_FONT
        r += 1
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 90
    wb.move_sheet("Summary", -(len(wb.sheetnames) - 1))


def build_report(db, excel_path: Path, reference: dict | None = None) -> dict:
    """Regenerate the workbook. Returns ``{sheets, numbered_missing}``."""
    excel_path = Path(excel_path)
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    ref_by_series = osu_api.reference_by_series(reference)
    cat_rows: dict[str, list] = defaultdict(list)
    numbered_missing: dict[str, list[int]] = {}

    for s in db.series_list():
        present = db.packs_for_series(s)
        category = present[0]["category"] if present else "Other"
        rows = gaps.build_rows(s, category, present, ref_by_series.get(s))
        cat_rows[category].extend(rows)
        miss = [r.number for r in rows if not r.present and r.number is not None]
        if miss:
            numbered_missing[s] = miss

    for p in db.packs_for_category("Other"):
        if p.get("series") is None:
            cat_rows["Other"].append(_present_row(None, p))

    for category in sorted(cat_rows):
        _category_sheet(wb, category, cat_rows[category])

    _artists_sheet(wb, db)
    _summary_sheet(wb, db, numbered_missing)

    if not wb.sheetnames:
        wb.create_sheet("Summary")
    wb.save(str(excel_path))
    return {"sheets": wb.sheetnames, "numbered_missing": numbered_missing}
